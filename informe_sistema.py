# Nombre del script: informe_sistema.py
# Propósito: recabar información de funcionamiento del sistema (almacenamiento usado, procesos que utilizan más memoria, etc) 
# y guardarlo en archivos .json y .xlsx
# Licencia: GPL v3 (www.gnu.org/licenses/gpl.html)

# HISTORIAL DE CAMBIOS
#
# Cuándo		Qué										Quién
# -----------	-------------------------------------	----------------------------------------	
# 09-Mar-2019   Versión inicial                         Gabriel Cánepa (Carrera Linux Argentina)
# 15-Mar-2019	Agregado reporte en formato Excel		Gabriel Cánepa (Carrera Linux Argentina)

import conversiones as cv
import json as j
import platform as p
import psutil as ps
from datetime import date
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

# Variables auxiliares
nombre_equipo = p.uname().node
fecha_actual = date.today().strftime('%Y%m%d')
reporte = nombre_equipo + '-' + fecha_actual

# Funciones auxiliares
def escribir_json(nombre_archivo, lista):
	with open(nombre_archivo, 'w') as archivo:
		j.dump(lista, archivo, indent = 4)

def llenar_tabla():
	pass
	
def dar_formato():
	pass
	
# Almacenamiento usado
uso_particiones = []
for particion in ps.disk_partitions():
	punto_montaje = particion.mountpoint
	total = ps.disk_usage(punto_montaje).total
	uso_particion = ps.disk_usage(punto_montaje).percent
	info_particion = {'punto_montaje': punto_montaje, 'total': cv.bytes2human(total), 'uso_particion': uso_particion}
	uso_particiones.append(info_particion)

# Escribir el contenido de uso_particiones al archivo particiones.json utilizando 4 espacios para la indentación
escribir_json('particiones.json', uso_particiones)

# Información sobre los 5 procesos que consumen más memoria
# Los campos que se pueden utilizar están disponibles en la documentación de psutil: https://psutil.readthedocs.io/en/latest/
info_procesos = []
atributos_proceso = ['pid', 'ppid', 'name', 'memory_percent', 'status', 'username']
for proceso in sorted(ps.process_iter(attrs = atributos_proceso), 
						key = lambda p: p.info['memory_percent'])[-5:]:
	info_procesos.append(proceso.info)
	
# Escribir el contenido de info_procesos al archivo procesos.json utilizando 4 espacios para la indentación
escribir_json('info_procesos.json', info_procesos)

# Crear planilla de cálculo
archivo_excel = Workbook()
archivo_excel.save('{}.xlsx'.format(reporte))

# El archivo se crea con una única pestaña, activa por defecto, a la que damos el título 'Particiones'
informe_particiones = archivo_excel.active
informe_particiones.title = 'Particiones'

# IMPORTANTE: Aunque el script en su forma actual funciona, las siguientes tres operaciones:
# 1) Insertar títulos de columnas,
# 2) Agregar datos de cada partición filas separadas, y
# 3) Dar formato a la tabla
# se repiten 2 veces (con ligeros cambios) al examinar la información sobre particiones primero y procesos luego.
# Se deja como tarea al alumno implementar más arriba las funciones (utilizando los parámetros necesarios):
# a) llenar_tabla
# b) dar_formato
# para evitar los problemas potenciales que resultan de duplicar código en un programa.

# Títulos de columnas
informe_particiones.append(['Punto de montaje', 'Espacio total', '% Uso'])

# Agregar datos de cada partición
for particion in uso_particiones:
	informe_particiones.append([
								particion['punto_montaje'], 
								particion['total'], 
								particion['uso_particion']
								])

# Identificar última celda ocupada y aplicar formato a toda la tabla								
ultima_celda = informe_particiones.cell(row = informe_particiones.max_row, column = informe_particiones.max_column).coordinate
tabla_particiones = Table(displayName = 'TablaParticiones', ref = 'A1:{}'.format(ultima_celda))	
estilo = TableStyleInfo(name = 'TableStyleMedium6', showRowStripes=True)
tabla_particiones.tableStyleInfo = estilo
informe_particiones.add_table(tabla_particiones)

# Crear nueva pestaña y activarla
informe_procesos = archivo_excel.create_sheet('Procesos')
archivo_excel.active = informe_procesos

# Títulos de columnas
informe_procesos.append(['PID', 'PPID', 'Proceso', '% Uso de memoria', 'Estado', 'Usuario'])

# Agregar datos de cada proceso
for proceso in info_procesos:
	informe_procesos.append([
							proceso['pid'],
							proceso['ppid'],
							proceso['name'],
							proceso['memory_percent'],
							proceso['status'],
							proceso['username']
							])

# Identificar última celda ocupada y aplicar formato a toda la tabla
ultima_celda = informe_procesos.cell(row = informe_procesos.max_row, column = informe_procesos.max_column).coordinate
tabla_procesos = Table(displayName = 'TablaProcesos', ref = 'A1:{}'.format(ultima_celda))	
estilo = TableStyleInfo(name = 'TableStyleMedium6', showRowStripes=True)
tabla_procesos.tableStyleInfo = estilo
informe_procesos.add_table(tabla_procesos)

# Guardar archivo
archivo_excel.save('{}.xlsx'.format(reporte))